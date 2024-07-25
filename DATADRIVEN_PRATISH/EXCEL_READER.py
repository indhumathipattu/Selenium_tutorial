import openpyxl

from DATADRIVEN_PRATISH.READ_EXCL import wb
from Datadriven.Read_Excel import sh


def getRowCount(path,sheetname):
    wb=openpyxl.load_workbook(path)
    sh1= wb[sheetname]
    rowcount=sh.max_row
    return rowcount

def getColCount(path, sheetName):
    we=openpyxl.load_workbook(path)
    sh1=wb[sheetName]
    return Colcount

def readdate(path,sheetname,rownum,columnum):
    wb = openpyxl.load_workbook(path)
    sh1=wb[sheetname]
    d1=sh.cell(row=rownum,column=coloumnum).value
    return d1

def writedate(path,sheetname,rownum,columnum):
    wb = openpyxl.load_workbook(path)
    sh1=wb[sheetname]
    d1 = sh.cell(row=rownum, column=coloumnum).value
    return d1