import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\mathi\\Desktop\\Book1.xlsx")
sh=wb['Sheet1']
print(sh.max_row)
print(sh.max_column)
date=sh.cell(row=1,column=1).value
print(date)
for i in range(1,sh.max_row+1):
    for j in range(1,sh.max_column+1):
        d1=sh.cell(row=i,column=j).value
        print(d1)
#line 5 mistake




