import openpyxl


def getRowCount(path, sheetName):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetName]
    rowcount = sh.max_row
    return rowcount


def getColCount(path, sheetName):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetName]
    colcount = sh.max_coloumn
    return colcount


def readData(path, sheetName, rowNum, colNum):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetName]
    data = sh.cell(row=rowNum, column=colNum).value
    return data


def writeData(path, sheetName, rowNum, colNum, writeData):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetName]
    sh.cell(row=rowNum,column=colNum).value=writeData
    wb.save(path)