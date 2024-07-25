import openpyxl

#load_workbook, # max_row, # Max_column, # value, #cell

wb = openpyxl.load_workbook("C:\\Users\\pc\\Desktop\\DDF.xlsx")
sh = wb['Sheet8']

# single cell write the data
#sh.cell(row=1,column=2).value='Python'

for i in range(1, 6):
    for j in range(1,4):
        sh.cell(row=i,column=j).value="Python"

wb.save("C:\\Users\\pc\\Desktop\\DDF.xlsx")
#row na said la irunthu count
#colum na mela irunthu varathu
