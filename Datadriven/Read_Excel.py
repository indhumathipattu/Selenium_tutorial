import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\pc\\Desktop\\DDF.xlsx")
sh = wb['Sheet7']
# to identify how many rows in the Excel sheet
print(sh.max_row)

# to identify how many columns in the Excel sheet
print(sh.max_column)

#to identify the single value in the excel sheet
data = sh.cell(row=4, column=2).value
print(data)

#to take all the values from the Excel sheet
print("============================")
for i in range(1,sh.max_row+1):
    for j in range(1, sh.max_column+1):
        d1 = sh.cell(row=i,column=j).value
        print(d1)